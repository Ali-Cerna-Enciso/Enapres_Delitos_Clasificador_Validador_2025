#!/usr/bin/env python3
"""
Script 8: Fusionador de Excel Original con Análisis
Combina datos originales con resultados del análisis y aplica formato.
"""

import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Agregar el directorio raíz del proyecto al path de Python
project_root = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(project_root))

try:
    from src.config.config_manager import config
except Exception as e:
    print(f"⚠️  Advertencia: No se pudo cargar config: {e}")
    config = None

# Obtener modo de procesamiento desde config
PROCESSING_MODE = config.get_processing_mode() if config else "complex"


class ExcelMerger:
    
    def __init__(self, dataset_name):
        self.dataset_name = dataset_name
        self.processing_mode = PROCESSING_MODE
        
        # Rutas de archivos
        self.original_excel = config.get_dataset_paths(dataset_name)['raw']
        self.analysis_excel = Path("data/results") / dataset_name / f"analisis_{dataset_name}.xlsx"
        self.output_excel = Path("data/results") / dataset_name / f"validado_{dataset_name}.xlsx"
        
        # Cargar configuración de procesamiento desde config
        proc_cfg = config.get_processing_config()
        
        # En COMPLEX: usar componentes de ID
        if self.processing_mode == "complex":
            self.id_components = proc_cfg.get('id_components', ['HOGAR', 'P201', 'ID', 'P424_ID'])
            self.id_column = 'ID_DELITO_IA'
            self.analysis_id_column = 'ID_DELITO_IA'
            self.columns_to_copy = [
                'MATCH_DELITO',
                'ESTADO_REVISION',
                'JUSTIFICACION_ACIERTO',
                'DETALLE_ERRORES_MODELO',
                'RAZONAMIENTO_GENERAL'
            ]
        else:  # SIMPLE: usar ID directo
            self.id_components = None
            self.id_column = proc_cfg.get('id_column', 'ID')
            self.analysis_id_column = 'ID_DELITO_IA'
            self.columns_to_copy = ['RAZONAMIENTO_GENERAL']
    
    def _get_available_filename(self, base_path: Path) -> Path:
        """
        Encuentra un nombre de archivo disponible.
        Si el archivo base está abierto/bloqueado, retorna versión numerada (_1, _2, etc.)
        
        Args:
            base_path: Ruta del archivo deseado (ej: validado_Test_Nuevo.xlsx)
        
        Returns:
            Path del archivo disponible para escritura
        """
        # Intentar con el nombre original primero
        if self._can_write_file(base_path):
            return base_path
        
        # Si está bloqueado, buscar versión numerada disponible
        stem = base_path.stem  # nombre sin extensión
        suffix = base_path.suffix  # extensión
        parent = base_path.parent
        
        version = 1
        while True:
            versioned_path = parent / f"{stem}_{version}{suffix}"
            if self._can_write_file(versioned_path):
                print(f"⚠️  Archivo '{base_path.name}' está abierto o bloqueado")
                print(f"   Guardando como: '{versioned_path.name}'")
                return versioned_path
            version += 1
            
            # Límite de seguridad para evitar bucle infinito
            if version > 100:
                raise RuntimeError(f"No se pudo encontrar nombre disponible después de 100 intentos")
    
    def _can_write_file(self, file_path: Path) -> bool:
        """
        Verifica si un archivo puede ser escrito (no está abierto/bloqueado).
        
        Args:
            file_path: Ruta del archivo a verificar
        
        Returns:
            True si el archivo puede ser escrito, False si está bloqueado
        """
        # Si el archivo no existe, está disponible
        if not file_path.exists():
            return True
        
        # Intentar abrir el archivo en modo escritura exclusivo
        try:
            # En Windows, intentar renombrar el archivo a sí mismo
            # Falla si está abierto en otra aplicación
            if os.name == 'nt':  # Windows
                os.rename(file_path, file_path)
            else:  # Linux/Mac
                # Intentar abrir en modo append
                with open(file_path, 'a'):
                    pass
            return True
        except (OSError, PermissionError):
            return False
    
    def load_original_data(self):
        if not self.original_excel.exists():
            raise FileNotFoundError(f"No se encuentra archivo original: {self.original_excel}")
        
        df = pd.read_excel(self.original_excel)
        
        # En modo COMPLEX: crear ID_DELITO_IA usando componentes
        if self.processing_mode == "complex":
            for col in self.id_components:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64').astype(str)
            
            df[self.id_column] = df[self.id_components].apply(
                lambda row: '-'.join(row.replace('<NA>', 'NULO')), axis=1
            )
        # En modo SIMPLE: el ID ya existe en la columna ID
        
        return df
    
    def load_analysis_data(self):
        if not self.analysis_excel.exists():
            raise FileNotFoundError(f"No se encuentra archivo de análisis: {self.analysis_excel}")
        
        df = pd.read_excel(self.analysis_excel)
        
        # Seleccionar solo las columnas necesarias
        columns_needed = [self.analysis_id_column] + self.columns_to_copy
        return df[columns_needed].copy()
    
    def merge_data(self, df_original, df_analysis):
        # Asegurar que las claves tengan el mismo tipo
        df_original[self.id_column] = df_original[self.id_column].astype(str)
        df_analysis[self.analysis_id_column] = df_analysis[self.analysis_id_column].astype(str)

        # Verificar qué columnas del análisis ya existen en el original
        original_columns = set(df_original.columns)
        existing_columns = []
        new_columns = []

        for col in self.columns_to_copy:
            if col in original_columns:
                existing_columns.append(col)
                print(f"✓ Columna '{col}' ya existe, se actualizarán los datos")
            else:
                new_columns.append(col)
                print(f"+ Columna '{col}' será agregada")

        # Realizar merge con sufijos para detectar columnas duplicadas
        df_merged = pd.merge(
            df_original,
            df_analysis,
            left_on=self.id_column,
            right_on=self.analysis_id_column,
            how='left',
            suffixes=('', '_new')
        )

        # Para las columnas que ya existían, actualizar con los nuevos valores
        for col in existing_columns:
            new_col_name = f"{col}_new"
            if new_col_name in df_merged.columns:
                # Actualizar valores donde hay datos del análisis (sobrescribir con nuevos datos)
                # Primero crear una copia de la columna original y luego combinar
                df_merged[col] = df_merged[new_col_name].where(df_merged[new_col_name].notna(), df_merged[col])
                # Eliminar la columna temporal con sufijo _new
                df_merged = df_merged.drop(columns=[new_col_name])

        if 'ESTADO_REVISION' in df_merged.columns:
            df_merged['ESTADO_REVISION'] = df_merged['ESTADO_REVISION'].fillna(4)

        if self.analysis_id_column != self.id_column:
            df_merged = df_merged.drop(columns=[self.analysis_id_column])
            
        return df_merged
    
    def apply_conditional_formatting(self, excel_path):
        # Definir colores
        fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        fill_gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Gris plomo
        
        # Cargar workbook y dataframe
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        df_temp = pd.read_excel(excel_path)
        
        # Encontrar columna MATCH_DELITO
        try:
            match_col_index = df_temp.columns.get_loc('MATCH_DELITO') + 1
        except KeyError:
            print("No se encontró columna MATCH_DELITO para formato")
            return
        
        # Aplicar colores y rellenar valores faltantes
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=match_col_index)
            
            if cell.value is None or pd.isna(cell.value):
                # Fila no procesada: colorear gris y marcar como "FALTA OBS."
                cell.fill = fill_gray
                cell.value = "FALTA OBS."
            elif cell.value == True:
                cell.fill = fill_green
            elif cell.value == False:
                cell.fill = fill_red
        
        workbook.save(excel_path)
    
    def generate_statistics(self, df_merged, df_analysis, original_ids):
        total_rows = len(df_merged)
        analysis_total = len(df_analysis)
        matches = df_merged['MATCH_DELITO'].notna().sum()
        successful_matches = df_merged['MATCH_DELITO'].sum() if matches > 0 else 0
        match_rate = (successful_matches / matches * 100) if matches > 0 else 0
        
        # Verificar qué registros del análisis no se encontraron en el original
        analysis_ids = set(df_analysis[self.analysis_id_column].astype(str))
        missing_in_original = analysis_ids - original_ids
        
        return {
            'total_rows': total_rows,
            'analysis_total': analysis_total,
            'records_with_analysis': matches,
            'successful_matches': successful_matches,
            'match_rate': match_rate,
            'coverage': (matches / total_rows * 100) if total_rows > 0 else 0,
            'missing_in_original': len(missing_in_original),
            'missing_ids': list(missing_in_original)[:10],  # Solo primeros 10 para el reporte
            'analysis_coverage': ((analysis_total - len(missing_in_original)) / analysis_total * 100) if analysis_total > 0 else 0
        }
    
    def process(self):
        # En modo SIMPLE, agregar columna RAZONAMIENTO_GENERAL al Excel original preservando formato
        if self.processing_mode == "simple":
            print(f"MODE: SIMPLE - Agregando RAZONAMIENTO_GENERAL al original (preservando formato)")
            
            # Cargar análisis con pandas para obtener mapeo ID -> Razonamiento
            df_analysis = self.load_analysis_data()
            razonamiento_map = dict(zip(df_analysis['ID_DELITO_IA'], df_analysis['RAZONAMIENTO_GENERAL']))
            
            # Cargar Excel original con openpyxl para preservar formato
            workbook = load_workbook(self.original_excel)
            sheet = workbook.active
            
            # Encontrar o crear columna RAZONAMIENTO_GENERAL
            # Buscar la columna ID para mapear
            id_col_index = None
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value == self.id_column:
                    id_col_index = col_idx
                    break
            
            if id_col_index is None:
                print(f"ERROR: No se encontró columna '{self.id_column}' en el Excel original")
                return {'error': 'ID column not found'}
            
            # Agregar columna RAZONAMIENTO_GENERAL al final
            last_col_idx = sheet.max_column + 1
            sheet.cell(row=1, column=last_col_idx, value='RAZONAMIENTO_GENERAL')
            
            # Llenar datos de razonamiento
            for row_idx in range(2, sheet.max_row + 1):
                id_value = sheet.cell(row=row_idx, column=id_col_index).value
                if id_value is not None:
                    # Intentar buscar con el tipo original (int o string según lo que tenga el Excel)
                    razonamiento = razonamiento_map.get(id_value, '')
                    # Si no encontró, intentar con string convertido
                    if not razonamiento:
                        id_value_str = str(id_value).strip()
                        razonamiento = razonamiento_map.get(id_value_str, '')
                    # Si aún no encontró, intentar con int
                    if not razonamiento:
                        try:
                            id_value_int = int(id_value)
                            razonamiento = razonamiento_map.get(id_value_int, '')
                        except (ValueError, TypeError):
                            pass
                    sheet.cell(row=row_idx, column=last_col_idx, value=razonamiento)
            
            # Guardar preservando formato
            self.output_excel.parent.mkdir(parents=True, exist_ok=True)
            output_file = self._get_available_filename(self.output_excel)
            workbook.save(output_file)
            
            stats = {
                'total_records': sheet.max_row - 1,
                'output_file': output_file,
                'mode': 'simple'
            }
            return stats
        
        # MODO COMPLEX: fusión completa
        df_original = self.load_original_data()
        df_analysis = self.load_analysis_data()
        original_ids = set(df_original[self.id_column].astype(str))
        
        df_merged = self.merge_data(df_original, df_analysis)
        
        self.output_excel.parent.mkdir(parents=True, exist_ok=True)
        
        # Obtener nombre de archivo disponible (con versionado si está bloqueado)
        output_file = self._get_available_filename(self.output_excel)
        
        df_merged.to_excel(output_file, index=False, engine='openpyxl')
        self.apply_conditional_formatting(output_file)
        
        stats = self.generate_statistics(df_merged, df_analysis, original_ids)
        stats['output_file'] = output_file
        return stats
    
    def print_report(self, stats):
        """Imprime reporte de estadísticas (diferente según modo)."""
        if stats.get('mode') == 'simple':
            # Modo SIMPLE: solo mostrar información básica
            print(f"✅ Archivo procesado (modo SIMPLE): {stats.get('total_records', 0)} registros")
            print(f"   Ubicación: {stats.get('output_file', 'desconocida')}")
        else:
            # Modo COMPLEX: mostrar concordancia completa
            if 'successful_matches' in stats:
                print(f"Concordancia: {stats['successful_matches']}/{stats['records_with_analysis']} ({stats['match_rate']:.1f}%)")

def main():
    dataset_name = config.default_dataset_name if config else None
    
    if not dataset_name:
        print("❌ ERROR: No se pudo obtener dataset desde config")
        sys.exit(1)
    
    results_dir = Path("data/results") / dataset_name
    if not results_dir.exists():
        print(f"❌ ERROR: Directorio de resultados no encontrado")
        print(f"   Buscando: {results_dir}")
        sys.exit(1)
    
    print("="*80)
    print("FUSIÓN CON EXCEL ORIGINAL")
    print("="*80)
    
    merger = ExcelMerger(dataset_name)
    
    # Mostrar registros originales
    df_original = merger.load_original_data()
    print(f"Registros originales: {len(df_original)}")
    
    # Procesar
    stats = merger.process()
    
    output_filename = stats['output_file'].name
    print(f"✅ Excel validado creado: {output_filename}")

if __name__ == "__main__":
    main()